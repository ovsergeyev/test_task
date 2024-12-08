from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from src.schemas.SFileStatus import SFileStatus
from src.schemas.SUser import SUser
from src.auth import get_current_user
from typing import Dict
import uuid
from datetime import datetime
import openpyxl
from io import BytesIO
import asyncio

router = APIRouter(prefix="/file_processing", tags=["File Processing"])

file_statuses: Dict[str, SFileStatus] = {}


@router.post("/upload_file")
async def upload_file(
    file: UploadFile = File(...), current_user: SUser = Depends(get_current_user)
):
    if not file.filename.endswith((".xls", ".xlsx")):
        raise HTTPException(status_code=400, detail="Invalid file format")

    file_id = str(uuid.uuid4())
    upload_datetime = datetime.now()
    file_statuses[file_id] = SFileStatus(
        upload_datetime=upload_datetime, status="uploaded"
    )

    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        asyncio.create_task(process_excel_file(file_id, workbook))
        return {"file_id": file_id, "status": file_statuses[file_id]}
    except Exception as e:
        file_statuses[file_id].status = "error"
        return {"file_id": file_id, "status": file_statuses[file_id]}


@router.get("/get_processing_status")
def get_processing_status(
    file_id: str, current_user: SUser = Depends(get_current_user)
) -> SFileStatus:
    if not file_id in file_statuses.keys():
        raise HTTPException(status_code=404, detail="Not found")
    return file_statuses[file_id]


async def process_excel_file(file_id, workbook):
    file_statuses[file_id].status = "processing"

    before_list = []
    after_list = []

    for sheet in workbook.worksheets:
        if sheet.max_row < 2:
            continue

        before_index = None
        after_index = None
        first_row = next(sheet.iter_rows(values_only=True))

        # определяем индексы колонок before и after учитывая что перед ними
        # не должно быть колонок с пустым заголовком
        for index, cell in enumerate(first_row):
            if cell is not None:
                if cell == "before":
                    before_index = index
                if cell == "after":
                    after_index = index
            else:
                break

        is_complete_before = False
        is_complete_after = False
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if before_index is not None and after_index is not None:
                if len(row) > before_index and not is_complete_before:
                    # при первом пустом значении в колонке она больше не обрабатывается
                    if row[before_index] is not None:
                        before_list.append(row[before_index])
                    else:
                        is_complete_before = True
                if len(row) > after_index and not is_complete_after:
                    # при первом пустом значении в колонке она больше не обрабатывается
                    if row[after_index] is not None:
                        after_list.append(row[after_index])
                    else:
                        is_complete_after = True

        # если колонки before и after найдены на листе, остальные листы не обрабатываются
        if before_index is not None and after_index is not None:
            break

    if before_list and after_list:
        set_before = set(before_list)
        set_after = set(after_list)
        difference = set_after.symmetric_difference(set_before)

        if len(difference) == 1:
            x = difference.pop()
            if len(set_before) < len(set_after):
                file_statuses[file_id].result = {"added": x}
            else:
                file_statuses[file_id].result = {"removed": x}
        # else:
        #     "Число X не определено"
    # else:
    #     "Не найдены необходимые колонки."

    processing_datetime = datetime.now()
    file_statuses[file_id].processing_datetime = processing_datetime
    file_statuses[file_id].status = "processing"
